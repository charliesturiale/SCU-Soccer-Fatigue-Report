import report_catapult
import report_vald
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
from sqlalchemy import create_engine, or_
from sqlalchemy.orm import Session
from models import Metric, Player, Roster, Team, PlayerMetricValue
from dotenv import load_dotenv
from composite_metrics import (
    compute_composite_metrics,
    COMPOSITE_FUNCS,
    get_composite_metric_name,
    get_required_metrics_for_composite,
    COMPOSITE_METRIC_METADATA
)
from datetime import datetime

load_dotenv()

# No need to do much in here yet, because building profiles just updates sql
def generate_report_handler(match_date: datetime = None):
    """
    Main handler for generating the full report.

    Args:
        match_date (datetime, optional): The date of the match to generate the report for.
                                         If None, defaults to the current date.
    """
    if match_date is None:
        match_date = datetime.now()

    # Get Catapult data and the identified reporting period
    catapult_report_df, report_period = report_catapult.get_catapult_report_metrics_main(match_date=match_date)

    # Determine the end date for the report to use for VALD data and the report label
    if report_period:
        report_end_date = report_period['end']
    else:
        # Fallback if no Catapult period is found
        report_end_date = match_date

    # Get VALD data, using the report's end date as the anchor for "recent" tests
    forcedecks_report_df, nordbord_report_df = report_vald.get_vald_report_metrics_main(match_date=report_end_date)

    print("Report generation complete")
    print(f"  Catapult data: {len(catapult_report_df) if catapult_report_df is not None else 0} players")
    print(f"  ForceDecks data: {len(forcedecks_report_df) if forcedecks_report_df is not None else 0} players")
    print(f"  NordBord data: {len(nordbord_report_df) if nordbord_report_df is not None else 0} players")

    create_report_table_and_export(catapult_report_df, forcedecks_report_df, nordbord_report_df, report_end_date.strftime("%m/%d/%Y"), "WSOC")
    return

def create_report_table_and_export(catapult_report_df, forcedecks_report_df, nordbord_report_df, report_date=None, team_name="WSOC"):
    """
    Compile metrics from all three dataframes into a formatted Excel report.

    Players from Catapult serve as the master list. Metrics from ForceDecks and NordBord
    are merged in based on player_name. Missing data shows as N/A.
    """

    # ============================================================================
    # CONFIGURATION SECTION - Modify these settings as needed
    # ============================================================================

    # Define which metrics to include in the report (use metric codes from the dataframe)
    SELECTED_METRICS = {
        'catapult': [
            'total_distance',               # 1. Total Distance
            'high_speed_distance',          # 2. High Speed Distance
            'percentage_max_velocity',      # 3. Percent Max Velocity
            'high_intensity_efforts',       # 4. High Intensity Efforts (derived)
            'player_load_per_minute',       # 5. Player Load Per Minute
        ],
        'forcedecks': [
            '6553698',  # 6. RSI-Modified
            '6553734', # 7. Concentric Impulse / BM
            '6553730',    # 8. Eccentric Decel Impulse / BM
        ],
        'nordbord': [
            'nordbord_strength_rel',  # 9. Bilateral Relative Strength (derived)
            'nordbord_asym',          # 10. Asymmetry Percentage (derived)
        ]
    }

    # Define which composite (z-score based) metrics to include
    # These will be computed AFTER z-scores are calculated for all metrics
    SELECTED_COMPOSITE_METRICS = [
        # 'explosive_output',     # REMOVED: Weighted combination of Peak Power/BM and Concentric Mean Force z-scores
    ]

    # Configure comparison type for each metric:
    # - 'reference': Compare to historical average (average_value in database)
    # - 'previous': Compare to previous week (previous_value in database)
    METRIC_COMPARISON_CONFIG = {
        # Catapult metrics - use reference (historical average)
        'total_distance': 'reference',
        'high_speed_distance': 'reference',
        'percentage_max_velocity': 'reference',
        'high_intensity_efforts': 'reference',
        'player_load_per_minute': 'reference',
        'total_player_load': 'reference',
        'gen2_acceleration_band6plus_total_effort_count': 'reference',

        # ForceDecks metrics - use reference (historical average)
        '6553607': 'reference',
        '6553698': 'reference',
        '6553604': 'reference',
        '6553619': 'reference',
        '6553734': 'reference',
        '6553730': 'reference',

        # NordBord metrics - use reference (historical average)
        'nordbord_strength_rel': 'reference',
        'nordbord_asym': 'reference',
        'leftMaxForce': 'reference',
        'rightMaxForce': 'reference',
        'leftAvgForce': 'reference',
        'rightAvgForce': 'reference',

        # Composite metrics - these are already z-scores, so comparison is N/A
        # But we need entries here for the formatting logic
        'explosiveness_index': 'reference',
        # 'explosive_output': 'reference',
    }

    # Configure z-score thresholds for conditional formatting
    # Defines what constitutes a "significant" deviation from the player's profile
    # Format: metric_code: threshold (absolute value of z-score)
    # Example: A threshold of 1.0 means values beyond ±1 standard deviation are significant
    Z_SCORE_THRESHOLDS = {
        # Catapult metrics
        'total_distance': 1.0,
        'high_speed_distance': 1.0,
        'percentage_max_velocity': 0.8,
        'high_intensity_efforts': 1.0,
        'player_load_per_minute': 1.0,

        # ForceDecks metrics
        '6553607': 0.75,  # Jump Height
        '6553698': 0.75,  # RSI-Modified
        '6553604': 0.75,  # Peak Power / BM
        '6553619': 0.75,  # Concentric Mean Force
        '6553734': 0.75,  # Concentric Impulse / BM
        '6553730': 0.75,  # Eccentric Decel Impulse / BM    

        # NordBord metrics
        'nordbord_strength_rel': 0.75,  # Bilateral Relative Strength
        'nordbord_asym': 0.5,          # Asymmetry Percentage

        # Composite metrics (already z-scores)
        'explosiveness_index': 1.25,
        # 'explosive_output': 1.25,
    }

    # ============================================================================
    # END CONFIGURATION
    # ============================================================================

    # Get metric names from database for proper column headers
    metric_metadata = get_metric_metadata_from_db()

    # Start with catapult data (master list of players)
    report_df = catapult_report_df[['player_name']].copy()

    # Track which metric codes are in which columns (for formatting later)
    column_to_metric_code = {}

    # Add selected Catapult metrics
    for metric_code in SELECTED_METRICS['catapult']:
        if metric_code in catapult_report_df.columns:
            # Use the friendly name from database if available
            column_name = metric_metadata.get(metric_code, {}).get("name", metric_code)
            report_df[column_name] = catapult_report_df[metric_code].round(2)
            column_to_metric_code[column_name] = metric_code

    # Merge ForceDecks data
    if not forcedecks_report_df.empty:
        for metric_code in SELECTED_METRICS['forcedecks']:
            if metric_code in forcedecks_report_df.columns:
                column_name = metric_metadata.get(metric_code, {}).get("name", metric_code)
                temp_df = forcedecks_report_df[['player_name', metric_code]].copy()

                # Apply scaling for RSI-Modified (cm -> m based calculation)
                if metric_code == '6553698':
                    temp_df[metric_code] = temp_df[metric_code] / 100.0

                temp_df.rename(columns={metric_code: column_name}, inplace=True)
                report_df = report_df.merge(
                    temp_df,
                    on='player_name',
                    how='left'
                )
                # Round to 2 decimal places
                if column_name in report_df.columns:
                    report_df[column_name] = report_df[column_name].round(2)
                    column_to_metric_code[column_name] = metric_code

    # Merge NordBord data
    if not nordbord_report_df.empty:
        for metric_code in SELECTED_METRICS['nordbord']:
            if metric_code in nordbord_report_df.columns:
                column_name = metric_metadata.get(metric_code, {}).get("name", metric_code)
                temp_df = nordbord_report_df[['player_name', metric_code]].copy()
                temp_df.rename(columns={metric_code: column_name}, inplace=True)
                report_df = report_df.merge(
                    temp_df,
                    on='player_name',
                    how='left'
                )
                # Round to 2 decimal places
                if column_name in report_df.columns:
                    report_df[column_name] = report_df[column_name].round(2)
                    column_to_metric_code[column_name] = metric_code

    # Get player positions and sort by position groups
    player_positions = get_player_positions()
    report_df = sort_players_by_position(report_df, player_positions)

    # Get reference values from database for conditional formatting and z-score calculation
    player_average_values = get_player_average_values()

    # Compute composite metrics (z-score based) AFTER all regular metrics are collected
    if SELECTED_COMPOSITE_METRICS:
        report_df = add_composite_metrics(
            report_df,
            player_average_values,
            column_to_metric_code,
            SELECTED_COMPOSITE_METRICS,
            METRIC_COMPARISON_CONFIG,
            forcedecks_report_df,
            nordbord_report_df
        )

    # Reorganize columns: player_name, position, then metrics
    # Get list of metric columns (everything except player_name and helper columns)
    helper_columns = ['position', 'position_group', 'position_group_name', 'is_separator']
    metric_columns = [col for col in report_df.columns if col not in ['player_name'] + helper_columns]

    # New column order: player_name, position, then metrics (position_group_name is not exported, used only for grouping)
    ordered_columns = ['player_name', 'position'] + metric_columns + ['position_group_name', 'position_group', 'is_separator']
    report_df = report_df[ordered_columns]

    # Load template workbook
    template_path = os.path.join(os.path.dirname(__file__), "Template.xlsx")
    if os.path.exists(template_path):
        # Load workbook - images should be preserved automatically by openpyxl
        wb = load_workbook(template_path)
        ws = wb.active
        print(f"Loaded template from {template_path}")

        # Debug: Check what visual elements exist in the template
        print(f"Template diagnostics:")
        print(f"  - Active sheet: {ws.title}")
        print(f"  - All sheets: {wb.sheetnames}")
        print(f"  - Images (_images): {len(ws._images) if hasattr(ws, '_images') else 0}")
        print(f"  - Charts (_charts): {len(ws._charts) if hasattr(ws, '_charts') else 0}")
        print(f"  - Drawings: {len(ws._drawing.twoCellAnchor) if hasattr(ws, '_drawing') and ws._drawing else 0}")

        # Check if there's a _rels part that might contain image relationships
        if hasattr(ws, '_rels'):
            print(f"  - Worksheet relationships: {len(ws._rels) if ws._rels else 0}")

        # Check for any formulas that might cause #VALUE! errors
        print(f"  - Checking for formulas in header area (rows 1-6):")
        for row in range(1, 7):
            for col in range(1, 20):  # Check first 20 columns
                cell = ws.cell(row=row, column=col)
                if cell.data_type == 'f':  # Formula
                    print(f"    Formula found at {cell.coordinate}: {cell.value}")

        # Unmerge cells, but preserve merged cells in the header area (rows 1-5) to preserve images
        # Only unmerge cells in row 6 and below where we'll be writing data
        merged_ranges = list(ws.merged_cells.ranges)
        print(f"  - Found {len(merged_ranges)} merged cell ranges")
        header_area_merged = []
        data_area_merged = []

        for merged_range in merged_ranges:
            # Check if any part of this merged range is in rows 1-5 (header area)
            if merged_range.min_row <= 5:
                header_area_merged.append(merged_range)
            else:
                data_area_merged.append(merged_range)

        print(f"  - Preserving {len(header_area_merged)} merged ranges in header area (rows 1-5)")
        print(f"  - Unmerging {len(data_area_merged)} merged ranges in data area (row 6+)")

        # Only unmerge cells in the data area
        for merged_range in data_area_merged:
            ws.unmerge_cells(str(merged_range))

        # After unmerging, check for and clear #VALUE! errors that might have been introduced
        print(f"  - Checking for #VALUE! errors after unmerging:")
        value_errors_cleared = []
        for row in range(1, 7):
            for col in range(1, 20):
                cell = ws.cell(row=row, column=col)
                if cell.value == '#VALUE!' or (isinstance(cell.value, str) and '#VALUE!' in str(cell.value)):
                    value_errors_cleared.append(cell.coordinate)
                    # Clear the #VALUE! error
                    cell.value = None
        if value_errors_cleared:
            print(f"    Found and cleared {len(value_errors_cleared)} #VALUE! errors in: {', '.join(value_errors_cleared)}")

        # Write date and team to cells B4 and B5 (without changing formatting)
        if report_date:
            ws['B4'].value = report_date
            print(f"  - Set B4 (Date) to: {report_date}")
        if team_name:
            ws['B5'].value = team_name
            print(f"  - Set B5 (Team) to: {team_name}")

        # Find the first empty row after template headers
        # Template has a large header section, start data at row 7
        template_header_rows = 6
        start_row = template_header_rows + 1
    else:
        print(f"Warning: Template not found at {template_path}, creating new workbook")
        wb = Workbook()
        ws = wb.active
        ws.title = "Match Report"
        start_row = 1

    # Get columns to export (exclude helper columns)
    export_columns = [col for col in report_df.columns if col not in ['is_separator', 'position_group', 'position_group_name']]

    print(f"\nDebug - Export columns ({len(export_columns)}): {export_columns}")

    # Create display names for headers
    header_display_names = []
    for col in export_columns:
        if col == 'player_name':
            header_display_names.append('Player Name')
        elif col == 'position':
            header_display_names.append('Pos')
        else:
            header_display_names.append(col)

    # Write headers to row 6 if using template, or row 1 if creating new workbook
    header_row = template_header_rows if start_row > 1 else 1

    if start_row == 1:
        # Creating new workbook - write headers to row 1
        ws.append(header_display_names)
        start_row = 2
    else:
        # Using template - write headers to row 6 (template_header_rows)
        for col_num, display_name in enumerate(header_display_names, 1):
            ws.cell(row=header_row, column=col_num, value=display_name)

    # Style the header row - only style columns we're actually using
    # Dark gray background with white text
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    # Enable text wrapping for headers
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    num_export_columns = len(export_columns)
    for col_num in range(1, num_export_columns + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows starting after template headers
    # Define gray fill for separator rows
    separator_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    separator_font = Font(bold=True, size=11)

    current_row = start_row
    for idx, row in report_df.iterrows():
        is_separator = row.get('is_separator', False)

        # Write each cell
        for col_num, col_name in enumerate(export_columns, 1):
            value = row[col_name]

            # For separator rows, write the position group name in the first column
            if is_separator and col_num == 1:
                value = row['position_group_name']

            cell = ws.cell(row=current_row, column=col_num, value=value)

            # Apply separator row formatting
            if is_separator:
                cell.fill = separator_fill
                cell.font = separator_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        current_row += 1

    # Apply styling to data cells - only to columns we're using
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Only iterate over the columns we actually wrote data to
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=num_export_columns), start=header_row):
        # Check if this is a separator row by looking at the dataframe
        df_row_idx = row_idx - start_row
        is_separator = False
        if 0 <= df_row_idx < len(report_df):
            is_separator = report_df.iloc[df_row_idx].get('is_separator', False)

        for cell in row:
            cell.border = thin_border

            # Skip header row styling (header is at header_row)
            if cell.row == header_row:
                continue

            # Skip template header rows (rows 1-5)
            if cell.row < header_row:
                continue

            # Don't override separator row formatting that was already applied
            if not is_separator:
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Handle N/A display for missing values (except in separator rows)
                if cell.value is None or (isinstance(cell.value, float) and pd.isna(cell.value)):
                    cell.value = "N/A"

    # Apply number formats to metric columns to show units
    for col_idx, header_name in enumerate(export_columns, 1):
        metric_code = column_to_metric_code.get(header_name)
        if not metric_code:
            continue

        # Get unit from DB metadata or composite metadata
        unit = metric_metadata.get(metric_code, {}).get("unit")
        if not unit and metric_code in COMPOSITE_METRIC_METADATA:
            unit = COMPOSITE_METRIC_METADATA[metric_code].get("unit")

        if not unit or unit == "ct" or unit == "z-score":  # Don't format counts or z-scores
            continue

        # Build the format string (e.g., 0.0" m" or 0.0"%")
        num_format = f'0.0" {unit}"' if unit != "%" else '0.0"%"'

        # Apply to all data cells in this column
        for row_idx in range(start_row, ws.max_row + 1):
            df_row_idx = row_idx - start_row
            if 0 <= df_row_idx < len(report_df):
                is_separator = report_df.iloc[df_row_idx].get('is_separator', False)
                if not is_separator:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Only apply if the cell contains a number
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = num_format

    # Don't auto-adjust column widths - preserve template widths
    # The template already has appropriate column widths configured
    pass

    # Debug: Print what metrics we're trying to format
    print(f"\nDebug - Column to Metric Code mapping:")
    for col_name, metric_code in column_to_metric_code.items():
        print(f"  {col_name} -> {metric_code}")

    print(f"\nDebug - Sample player average values (first player):")
    if player_average_values:
        first_player = list(player_average_values.keys())[0]
        print(f"  Player: {first_player}")
        for metric_code, values in player_average_values[first_player].items():
            print(f"    {metric_code}: reference={values['reference']}, std_dev={values['std_dev']}")

    # Apply conditional formatting based on background profile comparison
    apply_conditional_formatting(
        ws,
        report_df,
        player_average_values,
        column_to_metric_code,
        METRIC_COMPARISON_CONFIG,
        Z_SCORE_THRESHOLDS,
        SELECTED_COMPOSITE_METRICS
    )

    # Save the Excel file
    output_dir = "../output"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "match_report.xlsx")

    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"Match report exported successfully!")
    print(f"  Location: {output_path}")
    print(f"  Players: {len(report_df)}")
    print(f"  Metrics: {len(report_df.columns) - 1}")  # Exclude player_name
    print(f"{'='*60}\n")

    return


def add_composite_metrics(report_df, player_average_values, column_to_metric_code,
                         selected_composite_metrics, comparison_config,
                         forcedecks_report_df, nordbord_report_df):
    """
    Compute composite metrics based on z-scores of existing metrics.

    This function calculates z-scores for all metrics in the report, then uses those
    z-scores to compute composite metrics (e.g., explosiveness index from multiple
    jump/power metrics).

    Parameters
    ----------
    report_df : pd.DataFrame
        The report dataframe with player metrics
    player_average_values : dict
        Nested dict with reference/previous/std_dev values per player and metric
    column_to_metric_code : dict
        Maps column names to metric codes
    selected_composite_metrics : list
        List of composite metric codes to compute
    comparison_config : dict
        Configuration for which comparison type to use per metric
    forcedecks_report_df : pd.DataFrame
        Raw ForceDecks data (for accessing metrics not in the report)
    nordbord_report_df : pd.DataFrame
        Raw NordBord data (for accessing metrics not in the report)

    Returns
    -------
    pd.DataFrame
        Updated dataframe with composite metric columns added
    """
    if not player_average_values:
        print("Warning: No reference values available, skipping composite metrics")
        return report_df

    print(f"\nComputing {len(selected_composite_metrics)} composite metrics...")

    # For each player, compute z-scores for all metrics, then compute composite metrics
    composite_columns = {code: [] for code in selected_composite_metrics}

    for idx, row in report_df.iterrows():
        player_name = row['player_name']

        # Skip separator rows
        if not player_name or player_name == '':
            for code in selected_composite_metrics:
                composite_columns[code].append(None)
            continue

        # Get this player's reference values
        if player_name not in player_average_values:
            # No reference data for this player
            for code in selected_composite_metrics:
                composite_columns[code].append(None)
            continue

        player_metrics = player_average_values[player_name]

        # Build a dict of z-scores for all metrics (including those not in the report but needed for composites)
        z_scores = {}

        # First, compute z-scores for metrics already in the report
        for column_name, metric_code in column_to_metric_code.items():
            # Get the current value from the report
            current_value = row.get(column_name)

            # Skip if no data
            if current_value is None or pd.isna(current_value):
                continue

            # Get reference value and std dev
            if metric_code not in player_metrics:
                continue

            # Choose which reference value to use based on config
            comparison_type = comparison_config.get(metric_code, 'reference')
            if comparison_type == 'previous':
                average_value = player_metrics[metric_code]['previous']
            else:  # 'reference' or default
                average_value = player_metrics[metric_code]['reference']

            std_dev = player_metrics[metric_code]['std_dev']

            # Skip if no average value or std dev
            if average_value is None or std_dev is None or std_dev == 0:
                continue

            # Calculate z-score
            z_score = (current_value - average_value) / std_dev
            z_scores[metric_code] = z_score

        # Second, compute z-scores for any required component metrics not already in the report
        # This allows composite metrics to use metrics that aren't displayed in the final output
        for composite_code in selected_composite_metrics:
            required_metrics = get_required_metrics_for_composite(composite_code)
            for required_metric in required_metrics:
                # Skip if we already computed this z-score
                if required_metric in z_scores:
                    continue

                # Check if this metric exists in the player's profile
                if required_metric not in player_metrics:
                    continue

                # Try to get the current value from forcedecks_report_df or nordbord_report_df
                current_value = None
                if not forcedecks_report_df.empty and required_metric in forcedecks_report_df.columns:
                    player_row_fd = forcedecks_report_df[forcedecks_report_df['player_name'] == player_name]
                    if not player_row_fd.empty:
                        current_value = player_row_fd[required_metric].iloc[0]
                elif not nordbord_report_df.empty and required_metric in nordbord_report_df.columns:
                    player_row_nb = nordbord_report_df[nordbord_report_df['player_name'] == player_name]
                    if not player_row_nb.empty:
                        current_value = player_row_nb[required_metric].iloc[0]

                # Skip if no current value found
                if current_value is None or pd.isna(current_value):
                    continue

                # Get comparison type and reference values
                comparison_type = comparison_config.get(required_metric, 'reference')
                if comparison_type == 'previous':
                    average_value = player_metrics[required_metric]['previous']
                else:
                    average_value = player_metrics[required_metric]['reference']

                std_dev = player_metrics[required_metric]['std_dev']

                # Skip if no average value or std dev
                if average_value is None or std_dev is None or std_dev == 0:
                    continue

                # Calculate z-score
                z_score = (current_value - average_value) / std_dev
                z_scores[required_metric] = z_score

        # Compute composite metrics from z-scores
        composite_values = compute_composite_metrics(z_scores)

        # Add computed values to columns
        for code in selected_composite_metrics:
            value = composite_values.get(code)
            composite_columns[code].append(value)

    # Add composite metric columns to dataframe
    for code in selected_composite_metrics:
        column_name = get_composite_metric_name(code)
        report_df[column_name] = composite_columns[code]

        # Convert to numeric and round to 2 decimal places
        if column_name in report_df.columns:
            report_df[column_name] = pd.to_numeric(report_df[column_name], errors='coerce')
            report_df[column_name] = report_df[column_name].round(2)

        # Add to column mapping for formatting
        column_to_metric_code[column_name] = code

        print(f"  Added composite metric: {column_name} (code: {code})")

    return report_df


def get_metric_metadata_from_db():
    """
    Retrieve metric metadata from the database.

    Returns
    -------
    dict
        Dictionary mapping metric codes to metadata dicts
        Example: {'total_distance': {'name': 'Total Distance', 'unit': 'm'}, ...}
    """
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("Warning: DATABASE_URL not set, cannot load metric metadata")
        return {}

    try:
        engine = create_engine(db_url)
        with Session(engine) as session:
            # Get all metrics from database
            all_metrics = session.query(Metric).filter(or_(
                Metric.provider.like('catapult%'),
                Metric.provider.like('vald%'),
                Metric.provider.like('derived%')
            )).all()

            metric_map = {
                metric.code: {"name": metric.name, "unit": metric.unit}
                for metric in all_metrics
            }

            print(f"Loaded {len(metric_map)} metric names from database")
            return metric_map

    except Exception as e:
        print(f"Error loading metric names from database: {e}")
        return {}


def get_player_average_values():
    """
    Retrieve reference, previous, and std deviation values for all players from the database.

    Returns
    -------
    dict
        Nested dictionary: {player_name: {metric_code: {'reference': val, 'previous': val, 'std_dev': val}}}
    """
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("Warning: DATABASE_URL not set, cannot load reference values")
        return {}

    try:
        engine = create_engine(db_url)
        with Session(engine) as session:
            # Query all player metric values with player and metric info
            query = session.query(
                Player.first_name,
                Player.last_name,
                Metric.code,
                PlayerMetricValue.average_value,
                PlayerMetricValue.previous_value,
                PlayerMetricValue.std_deviation
            ).join(
                PlayerMetricValue, Player.id == PlayerMetricValue.player_id
            ).join(
                Metric, Metric.id == PlayerMetricValue.metric_id
            ).all()

            # Build nested dictionary
            player_values = {}
            for first_name, last_name, metric_code, avg_val, prev_val, std_val in query:
                player_name = f"{first_name} {last_name}".strip()

                if player_name not in player_values:
                    player_values[player_name] = {}

                player_values[player_name][metric_code] = {
                    'reference': float(avg_val) if avg_val is not None else None,
                    'previous': float(prev_val) if prev_val is not None else None,
                    'std_dev': float(std_val) if std_val is not None else None
                }

            print(f"Loaded reference values for {len(player_values)} players")
            return player_values

    except Exception as e:
        print(f"Error loading player reference values from database: {e}")
        return {}


def get_player_positions(team_name="WSOC"):
    """
    Retrieve player positions from the database via the Roster table.

    Parameters
    ----------
    team_name : str
        Name of the team to get positions for (default: "WSOC")

    Returns
    -------
    dict
        Dictionary mapping player_name to position
    """
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("Warning: DATABASE_URL not set, cannot load player positions")
        return {}

    try:
        engine = create_engine(db_url)
        with Session(engine) as session:
            # Query positions from Roster table (joined with Player and Team)
            query = session.query(
                Player.first_name,
                Player.last_name,
                Roster.position
            ).join(
                Roster, Player.id == Roster.player_id
            ).join(
                Team, Team.id == Roster.team_id
            ).filter(
                Team.name == team_name
            ).all()

            # Build dictionary
            player_positions = {}
            for first_name, last_name, position in query:
                player_name = f"{first_name} {last_name}".strip()
                player_positions[player_name] = position

            print(f"Loaded positions for {len(player_positions)} players on team {team_name}")
            return player_positions

    except Exception as e:
        print(f"Error loading player positions from database: {e}")
        return {}


def sort_players_by_position(report_df, player_positions):
    """
    Sort players by position group and add separator rows between groups.

    Position groups in order:
    1. GK (Goalkeepers)
    2. D (Defenders) - includes D, CB, OB
    3. M (Midfielders)
    4. F (Forwards)

    Parameters
    ----------
    report_df : pd.DataFrame
        DataFrame with player data
    player_positions : dict
        Dictionary mapping player names to positions

    Returns
    -------
    pd.DataFrame
        Sorted DataFrame with separator rows between position groups,
        includes 'position', 'position_group', and 'position_group_name' columns
    """
    # Define position group mapping and priority order
    POSITION_GROUPS = {
        'GK': 1,
        'D': 2, 'CB': 2, 'OB': 2,  # All defenders grouped together
        'M': 3,
        'F': 4
    }

    GROUP_NAMES = {1: 'GK', 2: 'D', 3: 'M', 4: 'F'}

    # Add position and group to dataframe
    report_df['position'] = report_df['player_name'].map(player_positions)
    report_df['position_group'] = report_df['position'].map(
        lambda pos: POSITION_GROUPS.get(pos, 999) if pos else 999
    )
    report_df['position_group_name'] = report_df['position_group'].map(
        lambda grp: GROUP_NAMES.get(grp, 'Unknown') if grp != 999 else 'Unknown'
    )

    # Sort by position group, then by player name within group
    report_df = report_df.sort_values(['position_group', 'player_name']).reset_index(drop=True)

    # Build result with separator rows between groups
    result_rows = []
    current_group = None

    for _, row in report_df.iterrows():
        group = row['position_group']
        group_name = row['position_group_name']

        # Add separator row when changing groups (except before first group)
        if current_group is not None and group != current_group:
            # Create separator row with group name
            separator = pd.Series({col: None for col in report_df.columns})
            separator['player_name'] = group_name  # Set position group name
            separator['position_group'] = group
            separator['position_group_name'] = group_name
            separator['is_separator'] = True
            result_rows.append(separator)

        # Mark regular rows as not separators
        row = row.copy()
        row['is_separator'] = False
        result_rows.append(row)
        current_group = group

    # Create new dataframe from rows
    result_df = pd.DataFrame(result_rows).reset_index(drop=True)

    # Log the grouping results
    print(f"\nPosition grouping:")
    for group_id in sorted(set(report_df['position_group'])):
        if group_id == 999:
            group_name = 'Unknown'
        else:
            group_name = GROUP_NAMES.get(group_id, 'Unknown')
        count = len(report_df[report_df['position_group'] == group_id])
        print(f"  {group_name}: {count} players")

    return result_df


def apply_conditional_formatting(worksheet, dataframe, player_average_values,
                                 column_to_metric_code, comparison_config, z_score_thresholds,
                                 composite_metrics):
    """
    Apply conditional formatting based on Z-scores relative to player's historical profile.

    Color coding uses a smooth gradient based on Z-score (standard deviations from mean):
    - Red spectrum: Above average (darker red = higher Z-score, more load/attention needed)
    - White: Near average (within threshold standard deviations)
    - Blue spectrum: Below average (darker blue = lower Z-score, less load/recovery)

    The gradient uses Z-scores capped at ±3 for visualization.
    Per-metric thresholds are configured in z_score_thresholds parameter.

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        The Excel worksheet to format
    dataframe : pd.DataFrame
        The report dataframe with player metrics
    player_average_values : dict
        Nested dict with reference/previous/std_dev values per player and metric
    column_to_metric_code : dict
        Maps column names to metric codes
    comparison_config : dict
        Configuration for which comparison type to use per metric
    z_score_thresholds : dict
        Per-metric z-score thresholds defining what's "significant" (white zone boundary)
    composite_metrics : list
        List of composite metric codes (these are already z-scores)
    """
    if not player_average_values:
        print("No reference values available for conditional formatting")
        return

    color_debug_count = [0]  # Mutable list to track debug prints

    def interpolate_color_zscore(z_score, white_threshold=1.0):
        """
        Create a smooth gradient from blue (low) → white (average) → red (high) based on Z-score.

        Parameters
        ----------
        z_score : float
            Z-score (number of standard deviations from mean)
        white_threshold : float
            Threshold for white zone (default 1.0 SD)

        Returns
        -------
        tuple : (PatternFill, Font)
            The fill color and font to apply
        """
        # Cap the Z-score for color calculation at ±3 standard deviations
        capped_z = max(-3, min(3, z_score))

        # Define color endpoints (RGB values)
        # Red for positive (higher than average): RGB(192, 80, 77) = #C0504D
        red = (192, 80, 77)
        # White for neutral: RGB(255, 255, 255) = #FFFFFF
        white = (255, 255, 255)
        # Blue for negative (lower than average): RGB(68, 114, 196) = #4472C4
        blue = (68, 114, 196)

        # Threshold around 0 to keep white (configurable by metric type)
        if abs(capped_z) <= white_threshold:
            r, g, b = white
            font_color = "000000"  # Black text
        elif capped_z > white_threshold:
            # Positive: interpolate from white to red
            # Normalize to 0-1 range (white_threshold to 3 maps to 0 to 1)
            ratio = min((capped_z - white_threshold) / (3.0 - white_threshold), 1.0)
            r = int(white[0] + (red[0] - white[0]) * ratio)
            g = int(white[1] + (red[1] - white[1]) * ratio)
            b = int(white[2] + (red[2] - white[2]) * ratio)
            # Use white text on darker reds
            font_color = "FFFFFF" if ratio > 0.5 else "000000"
        else:  # capped_z < -white_threshold
            # Negative: interpolate from white to blue
            # Normalize to 0-1 range (-white_threshold to -3 maps to 0 to 1)
            ratio = min((abs(capped_z) - white_threshold) / (3.0 - white_threshold), 1.0)
            r = int(white[0] + (blue[0] - white[0]) * ratio)
            g = int(white[1] + (blue[1] - white[1]) * ratio)
            b = int(white[2] + (blue[2] - white[2]) * ratio)
            # Use white text on darker blues
            font_color = "FFFFFF" if ratio > 0.5 else "000000"

        # Convert RGB to hex
        hex_color = f"{r:02X}{g:02X}{b:02X}"

        # Debug: Print first few color calculations with tight thresholds
        if white_threshold < 1.0 and color_debug_count[0] < 10:
            print(f"    Color calc: z={z_score:.2f}, threshold={white_threshold}, hex=#{hex_color}, white={abs(capped_z) <= white_threshold}")
            color_debug_count[0] += 1

        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        font = Font(color=font_color)

        return fill, font

    # Get column indices for each metric from the header row in the template
    # The template header is on row 6, data starts on row 7
    header_row = 6
    data_start_row = 7
    headers = [cell.value for cell in worksheet[header_row]]

    formatted_cells_count = 0
    skipped_reasons = {}

    for col_idx, column_name in enumerate(headers, 1):
        # Skip Player Name and Pos columns (display names)
        if column_name in ['Player Name', 'Pos', None]:
            continue

        # Get the metric code for this column
        metric_code = column_to_metric_code.get(column_name)
        if not metric_code:
            # Check if this is a derived or composite metric that might not be in the database
            if column_name not in skipped_reasons:
                skipped_reasons[column_name] = "No metric code mapping"
            continue

        # Get comparison type for this metric
        comparison_type = comparison_config.get(metric_code, 'reference')

        # Get z-score threshold for this metric (default to 1.0 if not configured)
        metric_threshold = z_score_thresholds.get(metric_code, 1.0)

        # Check if this is a composite metric (already a z-score)
        is_composite_metric = metric_code in composite_metrics

        # Iterate through each player row (starting from data_start_row)
        for row_idx in range(data_start_row, worksheet.max_row + 1):
            # Get player_name from column 1 (player_name is first column now)
            player_name_cell = worksheet.cell(row=row_idx, column=1)
            player_name = player_name_cell.value

            # Skip separator rows (they have position group names like 'GK', 'D', 'M', 'F')
            if player_name in ['GK', 'D', 'M', 'F', 'Unknown']:
                continue

            value_cell = worksheet.cell(row=row_idx, column=col_idx)
            current_value = value_cell.value

            # Skip if no data or N/A
            if current_value is None or current_value == "N/A" or not isinstance(current_value, (int, float)):
                continue

            # For composite metrics, the current value IS the z-score
            if is_composite_metric:
                z_score = current_value
                # Apply gradient color based on Z-score using configured threshold
                fill, font = interpolate_color_zscore(z_score, metric_threshold)
                value_cell.fill = fill
                value_cell.font = font
                formatted_cells_count += 1
                continue

            # For regular metrics, calculate z-score from reference values
            # Get reference value for this player and metric
            if player_name not in player_average_values:
                if column_name not in skipped_reasons:
                    skipped_reasons[column_name] = f"Player '{player_name}' not in average_values"
                continue

            player_metrics = player_average_values[player_name]
            if metric_code not in player_metrics:
                if column_name not in skipped_reasons:
                    skipped_reasons[column_name] = f"Metric code '{metric_code}' not in player metrics"
                continue

            # Choose which reference value to use based on config
            if comparison_type == 'previous':
                average_value = player_metrics[metric_code]['previous']
            else:  # 'reference' or default
                average_value = player_metrics[metric_code]['reference']

            # Get standard deviation for Z-score calculation
            std_dev = player_metrics[metric_code]['std_dev']

            # Skip if no average value or std dev
            if average_value is None or std_dev is None or std_dev == 0:
                if column_name not in skipped_reasons:
                    skipped_reasons[column_name] = f"Missing avg_value or std_dev (avg={average_value}, std={std_dev})"
                continue

            # Calculate Z-score: (current - mean) / std_dev
            z_score = (current_value - average_value) / std_dev

            # Apply gradient color based on Z-score using configured threshold
            fill, font = interpolate_color_zscore(z_score, metric_threshold)
            value_cell.fill = fill
            value_cell.font = font
            formatted_cells_count += 1

    print(f"\nApplied Z-score conditional formatting to {formatted_cells_count} cells across {len(column_to_metric_code)} metric columns")

    if skipped_reasons:
        print(f"\nDebug - Columns skipped from formatting:")
        for col_name, reason in skipped_reasons.items():
            print(f"  {col_name}: {reason}")


# RUN FILE
if __name__ == "__main__":
    generate_report_handler()